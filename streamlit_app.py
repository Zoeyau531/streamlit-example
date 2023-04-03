from collections import namedtuple
import altair as alt
import math
import pandas as pd
import streamlit as st

st.file_uploader(AnnualLeaveRecord.csv)

from datetime import datetime
from dateutil import relativedelta
import pandas as pd
import openpyxl
from datetime import date
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, Color
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment  

df0 = pd.DataFrame({"Name": ["Zoe", "A", "B", "C"],
                  "Join Date": ["02/07/2022", "10/08/2020", "12/03/2016", "12/03/2015"]})
df = df0.set_index("Name")


from openpyxl import Workbook

wb = openpyxl.load_workbook(r"C:\Users\Zoe\Desktop\Python Learn\Annual Leave\AnnualLeaveRecord.xlsx", read_only=False, data_only=True,keep_vba=True)



for EmployeeName in df.index:
    
    ws = wb.active

    if EmployeeName not in wb.sheetnames:
    
        wb.create_sheet(index=0, title = EmployeeName)
    
    
    wb.move_sheet(EmployeeName, -(len(wb.sheetnames)-1))
    index = (wb[EmployeeName])
                     
    index. merge_cells('B1:F1')
    index. merge_cells('G1:K1')
    cell1 = index.cell(row=1, column=2)
    cell2 = index.cell(row=1,column=7)  
    cell1.alignment = Alignment(horizontal='center', vertical='center')
    cell2.alignment = Alignment(horizontal='center', vertical='center')    
    
    index["A1"] = "Balance"
    index["A2"] = "Date (M/Y)"
    index["A3"] = "/"
    
    index["B1"] = "ANNUAL LEAVE"
    index["G1"] = "SICK LEAVE"
    
    index["B2"] = "Gained"
    index["C2"] = "Taken"
    index["D2"] = "Net"
    index["E2"] = "Cap"
    index["F2"] = "Accumulated AL"
    
    index["G2"] = "Gained"
    index["H2"] = "Taken"
    index["I2"] = "Net"
    index["J2"] = "Cap"
    index["K2"] = "Accumulated SL"
    row = index.row_dimensions[1]
    
    index["A1"].font = Font(color="00969696")
    index["A1"].font = Font(bold=True)
    index["A1"].fill = PatternFill("solid", fgColor="00969696")
    index["B1"].font = Font(bold=True)
    index["B1"].fill = PatternFill("solid", fgColor="00969696")
    index["C1"].font = Font(bold=True)
    index["C1"].fill = PatternFill("solid", fgColor="00969696")
    index["D1"].font = Font(bold=True)
    index["D1"].fill = PatternFill("solid", fgColor="00969696")
    index["E1"].font = Font(bold=True)
    index["E1"].fill = PatternFill("solid", fgColor="00969696")
    index["F1"].font = Font(bold=True)
    index["F1"].fill = PatternFill("solid", fgColor="00969696")
    index["G1"].font = Font(bold=True)
    index["G1"].fill = PatternFill("solid", fgColor="00969696")
    index["H1"].font = Font(bold=True)
    index["H1"].fill = PatternFill("solid", fgColor="00969696")
    index["I1"].font = Font(bold=True)
    index["I1"].fill = PatternFill("solid", fgColor="00969696")
    index["J1"].font = Font(bold=True)
    index["J1"].fill = PatternFill("solid", fgColor="00969696")
    index["K1"].font = Font(bold=True)
    index["K1"].fill = PatternFill("solid", fgColor="00969696")    
    index.column_dimensions['A'].width = 20
    index.column_dimensions['B'].width = 20
    index.column_dimensions['C'].width = 20
    index.column_dimensions['D'].width = 20
    index.column_dimensions['E'].width = 20
    index.column_dimensions['F'].width = 20
    index.column_dimensions['G'].width = 20
    index.column_dimensions['H'].width = 20
    index.column_dimensions['I'].width = 20
    index.column_dimensions['J'].width = 20
    index.column_dimensions['K'].width = 20
    
    index["A3"].fill = PatternFill("solid", fgColor="00969696")
    index["B3"].fill = PatternFill("solid", fgColor="00969696")
    index["C3"].fill = PatternFill("solid", fgColor="00969696")
    index["D3"].fill = PatternFill("solid", fgColor="00969696")
    index["E3"].fill = PatternFill("solid", fgColor="00969696")
    index["F3"] = 0
    index["G3"].fill = PatternFill("solid", fgColor="00969696")
    index["H3"].fill = PatternFill("solid", fgColor="00969696")
    index["I3"].fill = PatternFill("solid", fgColor="00969696")
    index["J3"].fill = PatternFill("solid", fgColor="00969696")
    index["K3"] = 0            

    print("-------------------------------------")
    print(EmployeeName)



    d1 = start_date_joined = print("Engaged Date(MM/DD/YYYY): ", df.loc[EmployeeName]["Join Date"])

    d1 = df.loc[EmployeeName]["Join Date"]


    today = datetime.today()
    d2 = today.strftime("%m/%d/%Y")
    print("Current Date(MM/DD/YYYY): ",d2)

    start_date = datetime.strptime(d1, "%m/%d/%Y")
    end_date = datetime.strptime(d2, "%m/%d/%Y")
    delta = relativedelta.relativedelta(end_date, start_date)
    num_months_joined = delta.months + (delta.years * 12)
    if delta.days > 15:
        num_months_joined = num_months_joined + 1

    print("Number of months engaged: ", num_months_joined)
    int(num_months_joined)
    Annual_Leaves_Count = 0
    int(Annual_Leaves_Count)
    if int(num_months_joined) <= 36:
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
    if int(num_months_joined) in range(37, 39):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 2
    if int(num_months_joined) in range(39, 49):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
    if int(num_months_joined) in range(49, 51):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 2
    if int(num_months_joined) in range(51, 61):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
    if int(num_months_joined) in range(61, 63):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 2
    if int(num_months_joined) in range(63, 73):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
    if int(num_months_joined) in range(73, 77):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 2
    if int(num_months_joined) in range(77, 85):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
    if int(num_months_joined) in range(85, 89):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 2
    if int(num_months_joined) in range(89, 97):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
    if int(num_months_joined) in range(97, 101):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 2
    if int(num_months_joined) in range(101, 109):
        Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
    if int(num_months_joined) >= 109:
        if int(num_months_joined) % 12 > 6:
            Annual_Leaves_Count = int(Annual_Leaves_Count) + 1
        if int(num_months_joined) % 12 <= 6:
            Annual_Leaves_Count = int(Annual_Leaves_Count) + 2
            
    print("Annual Leave Addition this month: ", Annual_Leaves_Count)
    
    Sick_Leaves_Count = 0
    int(Sick_Leaves_Count)
    if int(num_months_joined) <= 12:
        Sick_Leaves_Count = int(Sick_Leaves_Count) + 2
    if int(num_months_joined) > 12:
        Sick_Leaves_Count = int(Sick_Leaves_Count) + 4
        
    print("Sick Leave Addition this month: ", Sick_Leaves_Count)

    
######################################  Start Hide here when first established excel  #######################################

    max_row_for_a = max((a.row for a in index['A'] if a.value is not None))

    currentMonth = str(datetime.now().month)
    currentYear = str(datetime.now().year)

    new_row = str(max_row_for_a + 1)
    index["A"+new_row] = "As at "+currentMonth+"/"+currentYear
    index["B"+new_row] = Annual_Leaves_Count
    index["G"+new_row] = Sick_Leaves_Count

# AL net cap accumulate

    index["D"+new_row].value = int(index["B"+new_row].value - int((index["C"+new_row]).value))
     
    if (int(index["F"+str(max_row_for_a)].value) + int(index["D"+new_row].value)) <= 8:
        Accumulated_AL = int(index["F"+str(max_row_for_a)].value) + int(index["D"+new_row].value)
        index["E"+new_row] = "N/A"
    if int(index["F"+str(max_row_for_a)].value) + int(index["D"+new_row].value) > 8:
        Accumulated_AL = 8
        index["E"+new_row] = "CAPPED"
    
    index["F"+new_row] = Accumulated_AL

# SL net cap accumulate

    index["I"+new_row].value = int(index["G"+new_row].value - int((index["H"+new_row]).value))
     
    if (int(index["K"+str(max_row_for_a)].value) + int(index["I"+new_row].value)) <= 16:
        Accumulated_AL = int(index["K"+str(max_row_for_a)].value) + int(index["I"+new_row].value)
        index["J"+new_row] = "N/A"
    if int(index["K"+str(max_row_for_a)].value) + int(index["I"+new_row].value) > 16:
        Accumulated_AL = 16
        index["J"+new_row] = "CAPPED"
    
    index["K"+new_row] = Accumulated_AL    
    

    thin_border = Border(right=Side(style='thin'))
    for n in range(1, int(new_row) +1):
        index.cell(row=n,column=1).border = thin_border
    thick_border = Border(right=Side(style='thick'))
    for n in range(1, int(new_row) +1):
        index.cell(row=n,column=6).border = thick_border
    
    und_border = Border(bottom=Side(style='thick'))
    for n in range(1, int(index.max_column)+1):
        index.cell(row=2,column=n).border = und_border
    
    left_border = Border(left=Side(style='thin'))
    for n in range(1,int(new_row) +1):
        index.cell(row=n,column=12).border = left_border

######################################  End Hide here when first established excel  #######################################





wb.save(r"C:\Users\Zoe\Desktop\Python Learn\Annual Leave\AnnualLeaveRecord.xlsx")

wb.close()

st.write(Sick_Leaves_Count)
